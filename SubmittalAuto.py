import openpyxl

def get_maxrow(worksheet):

    max_col_row = 1
    empty_rows = 3

    for row in worksheet.iter_rows(min_row=1, max_row=3000):
        # Check if at least one cell in the row is not empty
       
        if row[0].value is not None:
            max_col_row += 1
    
    return max_col_row + empty_rows


def ltn(col_head):
    alph = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    #print(len(alph))
    return alph.index(col_head) + 1

Log_fp = "C:/Users/Simon.Arts/Desktop/PMT Submittal_Log_Automation_2024-07-23.xlsx"
Schedule_fp = "C:/Users/Simon.Arts/Desktop/Submittal Schedule.xlsx"
header_rows = 2

log_wb = openpyxl.load_workbook(Log_fp)
schedule_wb = openpyxl.load_workbook(Schedule_fp)

# Assuming the log data is in the first sheet (index 0) of log_wb
log_sheet = log_wb.worksheets[0]

# Assuming the schedule data is in the first sheet (index 0) of schedule_wb
schedule_sheet = schedule_wb.worksheets[0]
last_row_schedule = get_maxrow(schedule_sheet)

last_row_log = log_sheet.max_row

# Iterate through rows in the log sheet and copy to the schedule sheet

for row_index in range(1 + header_rows, last_row_log + 1):
    # Assuming you want to copy the entire row
    schedule_ri = last_row_schedule + row_index - header_rows
    #spec section 
    schedule_sheet.cell(row=schedule_ri, column=ltn('A'), value= log_sheet.cell(row=row_index, column=ltn('A')).value)

    #submittal cat
    schedule_sheet.cell(row=schedule_ri + row_index, column=ltn('D'), value= log_sheet.cell(row=row_index, column=ltn('E')).value)

    #Recieved 
    schedule_sheet.cell(row=schedule_ri + row_index, column=ltn('K'), value=log_sheet.cell(row=row_index, column=ltn('L')).value)

    #contractual due date
    schedule_sheet.cell(row=schedule_ri + row_index, column=ltn('L'), value=log_sheet.cell(row=row_index, column=ltn('O')).value)

    #Arup Response date
    schedule_sheet.cell(row=schedule_ri + row_index, column=ltn('P'), value=log_sheet.cell(row=row_index, column=ltn('Q')).value)

    #Arup Response
    schedule_sheet.cell(row=schedule_ri + row_index, column=ltn('O'), value=log_sheet.cell(row=row_index, column=ltn('W')).value)


    '''row_data = []
    for col_index in range(1, log_sheet.max_column + 1):
        cell_value = log_sheet.cell(row=row_index, column=col_index).value
        row_data.append(cell_value)'''
    
    # Append the row to the schedule sheet


# Save changes to the schedule workbook
schedule_wb.save(Schedule_fp)

# Close the workbooks
log_wb.close()
schedule_wb.close()

print("Rows copied successfully from log workbook to schedule workbook.")